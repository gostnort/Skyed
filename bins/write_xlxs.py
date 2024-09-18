from openpyxl.utils import get_column_letter, column_index_from_string
import yaml
from openpyxl import load_workbook
from io import BytesIO

class FillOut():
    def __init__(self, ResourcePath):
        self.__path = ResourcePath
        with open(self.__path + r'\excel.yml', 'r') as file:
            yaml_structure = yaml.safe_load(file)
        self.__yaml = yaml_structure
        template_path = self.__yaml['excel_template_path']
        # Load the template file into memory
        with open(template_path, 'rb') as template_file:
            template_data = template_file.read()

        # Create a BytesIO object from the template data
        template_in_memory = BytesIO(template_data)

        # Load the workbook from the BytesIO object
        self.__Workbook = load_workbook(template_in_memory, keep_vba=False)

        # Select the active sheet (or specify a sheet name)
        self.WorkSheet = self.__Workbook.active


    def __get_cells(self,cell_range):
        cells = []
        
        # Split the input to get start and end cell references
        if ':' in cell_range:
            start_cell, end_cell = cell_range.split(':')
        else:
            start_cell = end_cell = cell_range

        # Extract column letters and row numbers from the cell references
        start_col_letter = ''.join([char for char in start_cell if char.isalpha()])
        start_row = ''.join([char for char in start_cell if char.isdigit()])

        end_col_letter = ''.join([char for char in end_cell if char.isalpha()])
        end_row = ''.join([char for char in end_cell if char.isdigit()])

        # Convert column letters to indices
        start_col_index = column_index_from_string(start_col_letter)
        end_col_index = column_index_from_string(end_col_letter)

        # Generate cell references for the entire range
        for col_index in range(start_col_index, end_col_index + 1):
            for row_index in range(int(start_row), int(end_row) + 1):
                col_letter = get_column_letter(col_index)
                cells.append(f"{col_letter}{row_index}")
        
        return cells

    def WriteArrivalFlight(self, ArrivalFlight, Bol_Left=False):
        if Bol_Left:
            cell_name = self.__yaml['left_side']['arrival_flight_number']
        else:
            cell_name = self.__yaml['right_side']['arrival_flight_number']
        self.WorkSheet[cell_name] = ArrivalFlight

    def WriteArrivalLeg(self, ArrivalLeg, Bol_Left=False):
        if Bol_Left:
            cell_name = self.__yaml['left_side']['arrival_leg']
        else:
            cell_name = self.__yaml['right_side']['arrival_leg']
        self.WorkSheet[cell_name] = ArrivalLeg[:3]+'-'+ArrivalLeg[3:]

    def WriteDepartureFlight(self, DepartureFlight, Bol_Left=False):
        if Bol_Left:
            cell_name = self.__yaml['left_side']['departure_flight_number']
        else:
            cell_name = self.__yaml['right_side']['departure_flight_number']
        self.WorkSheet[cell_name] = DepartureFlight

    def save_copy(self, FileName):
        output_in_memory = BytesIO()
        self.__Workbook.save(output_in_memory)
        # Seek to the start of the BytesIO object before reading or writing
        output_in_memory.seek(0)
        # Save the content of the BytesIO object to a new .xlsx file
        with open(self.__path + '\\' + FileName, 'wb') as new_file:
            new_file.write(output_in_memory.getbuffer())
        # Make sure the workbook is closed
        self.__Workbook.close()

    def WriteDepartureLeg(self, DepartureLeg, Bol_Left=False):
        if Bol_Left:
            cell_name = self.__yaml['left_side']['departure_leg']
        else:
            cell_name = self.__yaml['right_side']['departure_leg']
        self.WorkSheet[cell_name] = DepartureLeg[:3]+'-'+DepartureLeg[3:]

    def WriteDepartureDate(self, DepartureDate, Bol_Left=False):
        if Bol_Left:
            cell_name = self.__yaml['left_side']['departure_flight_date']
        else:
            cell_name = self.__yaml['right_side']['departure_flight_date']
        self.WorkSheet[cell_name] = DepartureDate

    def WriteArrivalSeatConfiguration(self, SeatCnf, Bol_Left=False):
        if Bol_Left:
            cell_name = self.__yaml['left_side']['arrival_seat_configuration']
        else:
            cell_name = self.__yaml['right_side']['arrival_seat_configuration']
        self.WorkSheet[cell_name] = SeatCnf

    def WriteDepartureEtd(self, DepartureEtd, Bol_Left=False):
        if Bol_Left:
            cell_name = self.__yaml['left_side']['departure_std']
        else:
            cell_name = self.__yaml['right_side']['departure_std']
        self.WorkSheet[cell_name] = 'STD:' + DepartureEtd
        if Bol_Left:
            cell_name = self.__yaml['left_side']['departure_etd']
        else:
            cell_name = self.__yaml['right_side']['departure_etd']
        self.WorkSheet[cell_name] = 'ETD:' + DepartureEtd

    def WriteDepartureEta(self, DepartureEta, Bol_Left=False):
        if Bol_Left:
            cell_name = self.__yaml['left_side']['departure_eta']
        else:
            cell_name = self.__yaml['right_side']['departure_eta']
        self.WorkSheet[cell_name] = 'ETA:' + DepartureEta
    
    def WriteArricalAc(self, AcReg, Bol_Left=False):
        if Bol_Left:
            cell_name = self.__yaml['left_side']['arrival_ac_reg']
        else:
            cell_name = self.__yaml['right_side']['arrival_ac_reg']
        self.WorkSheet[cell_name] = AcReg

    def WriteDepartureBdt(self, DepartureBdt, Bol_Left=False):
        if Bol_Left:
            cell_name = self.__yaml['left_side']['departure_boarding_time']
        else:
            cell_name = self.__yaml['right_side']['departure_boarding_time']
        self.WorkSheet[cell_name] = DepartureBdt

    def WriteArrivalPax(self, ArrivalPax, Bol_Left=False):
        if Bol_Left:
            cell_name = self.__yaml['left_side']['arrival_pax_break_down']
        else:
            cell_name = self.__yaml['right_side']['arrival_pax_break_down']
        self.WorkSheet[cell_name] = ArrivalPax

    def WriteDeparturePax(self, DeparturePax, Bol_Left=False):
        if Bol_Left:
            cell_name = self.__yaml['left_side']['departure_pax_break_down']
        else:
            cell_name = self.__yaml['right_side']['departure_pax_break_down']
        self.WorkSheet[cell_name] = DeparturePax

    def WriteDepartureGate(self, Gate, Bol_Left=False):
        if Bol_Left:
            cell_name = self.__yaml['left_side']['departure_gate']
        else:
            cell_name = self.__yaml['right_side']['departure_gate']
        self.WorkSheet[cell_name] = Gate

    def WriteSpecials(self, Specials, Bol_Left=False):
        if Bol_Left:
            cells_range = self.__yaml['left_side']['Special_count']
        else:
            cells_range = self.__yaml['right_side']['Special_count']
        cells = self.__get_cells(cells_range)
        for cell_tmp, key in zip(cells, Specials):
            self.WorkSheet[cell_tmp] = key + ' ' + str(Specials[key])

    def WriteComments(self, Comments, Bol_Left=False):
        if Bol_Left:
            cells_range = self.__yaml['left_side']['Comment']
        else:
            cells_range = self.__yaml['right_side']['Comment']
        cells = self.__get_cells(cells_range)
        for cell_tmp, key in zip(cells, Comments):
            self.WorkSheet[cell_tmp] = Comments[key]

def main():
    xls = FillOut(r'C:\Users\gostn\my_github\skyed\resources')
    xls.WriteArrivalFlight("CA111")
    xls.WriteArrivalLeg('SZX-LAX')
    xls.save_copy('new_sheet.xlsx')

if __name__ == "__main__":
    main()